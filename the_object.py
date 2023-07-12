from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter

from main import Facade, Ceiling


class TheObject:
    def __init__(self, *args):
        self.facades = []
        self.ceilings = []
        for el in args:
            if type(el) == Facade:
                self.facades.append(el)
            elif isinstance(el, Ceiling):
                self.ceilings.append(el)
            else:
                raise ValueError(f"{el} не относится к {Facade} или {Ceiling}")

    @property
    def _everything(self):
        lst = [el for el in self.facades]
        lst.extend(self.ceilings)
        return lst

    def square(self, facade=True, ceiling=True, ral=None):
        total = TheObject._square(self.facades, ral) if facade else 0
        total += TheObject._square(self.ceilings, ral) if ceiling else 0
        return total

    @staticmethod
    def _square(lst: list[Facade | Ceiling], ral):
        return round(sum([el.square(ral) for el in lst]), 2)

    def get_list_in_production(self, the_date=None, filename=None):
        lst = []
        for area in self._everything:
            lst.extend(area.get_list_in_production(the_date))
        counter = Counter(lst)
        if filename:
            TheObject._create_xlsx_from_counter(filename, counter)
        return counter

    def get_list_cassette_by_ral_or_size(self, width=None, height=None, ral=None, filename=None):
        lst = []
        for area in self._everything:
            lst.extend(area.get_list_cassette_by_ral_or_size(width, height, ral))
        counter = Counter(lst)
        if filename:
            TheObject._create_xlsx_from_counter(filename, counter)
        return counter

    @staticmethod
    def _create_xlsx_from_counter(filename, counter):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Заполняем заголовки столбцов
        sheet['A1'] = '№п/п'
        sheet['B1'] = 'Элемент'
        sheet['C1'] = 'Количество'

        # Заполняем данные из collections.Counter
        row = 2
        number = 1
        for item, count in sorted(counter.items(), key=lambda x: (-x[1], x[0])):
            sheet[f'A{row}'] = number
            sheet[f'B{row}'] = str(item)
            sheet[f'C{row}'] = count
            row += 1
            number += 1

        # Авторазмер столбцов
        for column_cells in sheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)  # Получаем имя столбца
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filename)
