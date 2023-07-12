from openpyxl.utils import get_column_letter
from copy import deepcopy

from main import Cassette
from collections import Counter
import openpyxl
from novoros import novoros


class CassettesList:
    def __init__(self, cassettes: Counter):
        cas = deepcopy(cassettes)
        self.cassettes = cas
        self.cassettes.most_common()

    def __repr__(self):
        return f"CassettesList({self.cassettes})"

    def __add__(self, other):
        if isinstance(other, CassettesList):
            cassettes = deepcopy(self.cassettes)
            cassettes.update(other.cassettes)
            return CassettesList(cassettes)
        raise NotImplemented

    def __sub__(self, other):
        if isinstance(other, CassettesList):
            cassettes = deepcopy(self.cassettes)
            cassettes.subtract(other.cassettes)
            return CassettesList(cassettes)
        raise NotImplemented

    def create_xlsx_from_counter(self, filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Заполняем заголовки столбцов
        sheet['A1'] = '№п/п'
        sheet['B1'] = 'Элемент'
        sheet['C1'] = 'Количество'

        # Заполняем данные из collections.Counter
        row = 2
        number = 1
        for item, count in self.cassettes.items():
            if count:
                sheet[f'A{row}'] = number
                sheet[f'B{row}'] = str(item)
                sheet[f'C{row}'] = count
                number += 1
                row += 1

        # Авторазмер столбцов
        for column_cells in sheet.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)  # Получаем имя столбца
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(filename)


def list_of_tuple_from_xlsx(filename, *rals):
    result_dict = {}

    # Открываем файл
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Проходим по строкам
    for row in ws.iter_rows():
        # Проверяем значение столбца
        if row[0].value in rals:
            # Добавляем кортеж в результирующий список
            cas = Cassette(int(round(row[6].value, 0)), row[7].value, row[0].value)
            number = row[10].value
            if cas in result_dict:
                result_dict[cas] = number + result_dict[cas]
            else:
                result_dict[cas] = number

    # Закрываем файл
    wb.close()

    # Возвращаем результирующий список
    return result_dict


cas1 = Counter(list_of_tuple_from_xlsx("data/cassette_origin.xlsx", 9006))
print("cas1 = ", cas1)
caslist = CassettesList(cas1)
print(caslist)
caslist2 = CassettesList(novoros.get_list_cassette_by_ral_or_size(ral=9006))
print("novoros: ", caslist2)
total = caslist + caslist2
print("total = ", total)
sub = caslist - caslist2

sub.create_xlsx_from_counter("data/sub_class.xlsx")
print("sub = ", sub)
